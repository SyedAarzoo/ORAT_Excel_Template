import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
# Read the Excel file into a pandas dataframe
df = pd.read_excel('19AJ100429-GC3-FM-024-A1-TRENCHER-DATA-SL_T1500_Dive_025_2nd_Pass.xlsx', sheet_name='T1500', engine='openpyxl')

# Get the indices of the blank rows
blank_row_indices = df.index[df.isnull().all(axis=1)]

file_cnt = 0
# Loop through the indices and create new dataframes for each section
for i in range(len(blank_row_indices)):
    if i == 0:
        section_df = df.loc[:blank_row_indices[i]]
    else:
        section_df = df.loc[blank_row_indices[i-1]+1:blank_row_indices[i]]
    
    # Create a new Excel file for the section
    section_df.to_excel(f'Section{i+1}.xlsx', index=False)
    file_cnt += 1
print(file_cnt)
section_df = df.loc[blank_row_indices[-1]+1:]
# Create a new Excel file for the section
section_df.to_excel(f'Section{len(blank_row_indices)+1}.xlsx', index=False)


for i in range(0,file_cnt+1):

    # Load the source workbook and sheet
    source_wb = openpyxl.load_workbook(f'Section{i+1}.xlsx')
    source_ws = source_wb['Sheet1']
    source_header = [cell.value for cell in source_ws[1]]

    # Load the destination workbook and sheet
    dest_wb = openpyxl.load_workbook('SHET_GC3-T1500_FM-XXX_Dive_XXX_EMPTY.xlsx')
    dest_ws = dest_wb['Data']
    dest_header = [cell.value for cell in dest_ws[2]]

    df = pd.DataFrame(list(source_ws.values))
    df.columns = df.iloc[0,:]
    df = df.iloc[1:,].reset_index(drop=True)

    df1 = pd.DataFrame(list(dest_ws.values))
    df1.columns = df1.iloc[1,:]
    df1 = df1.iloc[2:,].reset_index(drop=True)

    df['Port Fwd Load\n[bar]'] = df['Port Fwd Load\n[bar]'].astype(float)
    df['Stbd Fwd Load\n[bar]'] = df['Stbd Fwd Load\n[bar]'].astype(float)

    df['Average'] = (df['Port Fwd Load\n[bar]'] + df['Stbd Fwd Load\n[bar]'])/2

    df1['Tool Fwd/Aft Load \n[bar]'] = df['Average']

    df1[['Date\n[YYYY-MM-DD]', 'Time\n[HH:MM:SS]','KP\n[m]','DCC\n[m]','Depth\n[m]','Pitch\n[d.ddº]','Roll\n[d.ddº]','Port Sword\n[m]','Stbd Sword\n[m]','Jet Water Pressure\n[bar]','Comment']] = \
    df[["Date\n[YYYY-MM-DD]", "Time\n[HH:MM:SS]",'KP\n[km]','DCC\n[m]','Depth\n[m]','Pitch\n[d.ddº]','Roll\n[d.ddº]','Port Tip Depth\n[m]','Stbd Tip Depth\n[m]', 'Jet Water Pressure \n[bar]','Comment']]
    #df1[[df1.columns[5],df1.columns[6]]] = df[[df.columns[4],df.columns[5]]]

    try:
        df1[['Easting\n[m]','Northing\n[m]']]=df[['Easting\n[m]','Northing\n[m]','Easting\n[m]','Northing\n[m]']]
        
    except:
        df1[['Easting\n[m]', 'Northing\n[m]']] = df[['Easting\n[m]', 'Northing\n[m]']]
    

        
    col_indices = [2,3,4]
    df1.iloc[:, col_indices] = None

    rows = dataframe_to_rows(df1, index=False)

    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            dest_ws.cell(row=r_idx, column=c_idx, value=value)

    df1.to_excel(f'SHET_GC3-T1500_FM-XXX_Dive_XXX_EMPTYSection{i+1}.xlsx', index=False)

    # Save the destination workbook
    dest_wb.save(f'SHET_GC3-T1500_FM-XXX_Dive_XXX_EMPTYSection{i+1}.xlsx')
