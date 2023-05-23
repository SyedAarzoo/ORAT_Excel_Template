import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Load the source workbook and sheet
source_wb = openpyxl.load_workbook('Section1.xlsx')
source_ws = source_wb['Sheet1']
source_header = [cell.value for cell in source_ws[1]]

# Load the destination workbook and sheet
dest_wb = openpyxl.load_workbook('SHET_GC3-T1500_FM-XXX_Dive_XXX_EMPTY.xlsx')
dest_ws = dest_wb['Data']
dest_header = [cell.value for cell in dest_ws[2]]

# # Iterate over the columns and print the header
# for col in dest_ws.iter_cols(1, dest_ws.max_column):
#     header = col[0].value
#     print(header)

df = pd.DataFrame(list(source_ws.values))
df.columns = df.iloc[0,:]
df = df.iloc[1:,].reset_index(drop=True)

#print(df.columns)

df1 = pd.DataFrame(list(dest_ws.values))
df1.columns = df1.iloc[1,:]
df1 = df1.iloc[2:,].reset_index(drop=True)

#print(df1.columns)
#print(df.columns)
#column_id = df.columns.get_loc('Easting\n[m]')


# change = []
# for header in dest_header:
#     if header in source_header:
#         change.append(header)

# Convert column data from string to float
df['Port Fwd Load\n[bar]'] = df['Port Fwd Load\n[bar]'].astype(float)
df['Stbd Fwd Load\n[bar]'] = df['Stbd Fwd Load\n[bar]'].astype(float)

# Perform mathematical operations
df['Average'] = (df['Port Fwd Load\n[bar]'] + df['Stbd Fwd Load\n[bar]'])/2

# Copy the result to another column
df1['Tool Fwd/Aft Load \n[bar]'] = df['Average']


df1[['Date\n[YYYY-MM-DD]', 'Time\n[HH:MM:SS]','KP\n[m]','DCC\n[m]','Depth\n[m]','Pitch\n[d.ddº]','Roll\n[d.ddº]','Port Sword\n[m]','Stbd Sword\n[m]','Jet Water Pressure\n[bar]','Comment']] = \
df[["Date\n[YYYY-MM-DD]", "Time\n[HH:MM:SS]",'KP\n[km]','DCC\n[m]','Depth\n[m]','Pitch\n[d.ddº]','Roll\n[d.ddº]','Port Tip Depth\n[m]','Stbd Tip Depth\n[m]', 'Jet Water Pressure \n[bar]','Comment']]
#df1[[df1.columns[5],df1.columns[6]]] = df[[df.columns[4],df.columns[5]]]

df1[['Easting\n[m]','Northing\n[m]']]=df[['Easting\n[m]', 'Northing\n[m]','Easting\n[m]','Northing\n[m]']]

col_indices = [2,3,4]   
df1.iloc[:, col_indices] = None

#df1 = df 
#df1[change] = df[change]

#print(df1.head(1).transpose())

rows = dataframe_to_rows(df1, index=False)

for r_idx, row in enumerate(rows, 2):
    for c_idx, value in enumerate(row, 1):
        
        dest_ws.cell(row=r_idx, column=c_idx, value=value)
     




df1.to_excel('SHET_GC3-T1500_FM-XXX_Dive_XXX_EMPTYsec1.xlsx', index=False)

# Save the destination workbook
dest_wb.save('SHET_GC3-T1500_FM-XXX_Dive_XXX_EMPTYsec1.xlsx')


