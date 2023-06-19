import math
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import os


folder_path = r'C:\Users\ASyed\Downloads\ORAT Template Extraction Scripts'  # Replace with the path to your folder

xlsx_files = []



for file in os.listdir(folder_path):
    if file.endswith('.xlsx'):
        xlsx_files.append(file)

for file in xlsx_files:
    try:
        # Read the Excel file into a pandas dataframe
        df = pd.read_excel(file, sheet_name='T1500', engine='openpyxl')
        

    except KeyError:
        
        pass
    
    except ValueError:
        pass
    else:
        df_name = pd.read_excel(file, sheet_name='Additional Info', engine='openpyxl', skiprows=28)
        dive_no = df_name["Dive No"][0].split(" ")[1].strip('#')
        
        df_FM = pd.read_excel(file, sheet_name='Cover', engine='openpyxl', skiprows=16)
        FM_col = df_FM.columns
        FM_no = FM_col[2].split("-")[3]
        kp_from = list(df_name["KP From"])
        kp_to = list(df_name["KP To"])
        kp2 = [x for x in kp_from if not math.isnan(x)]
        kp1 = [x for x in kp_to if not math.isnan(x)]
        kp1 = sorted(kp1)
        kp2 = sorted(kp2)
        
        # Check if the dataframe contains any blank rows
        if df.index[df.isnull().all(axis=1)].any():
            # Script 1: If the Excel file contains blank rows
            # Script 1 code here
            # ...
            # Get the indices of the blank rows
            blank_row_indices = df.index[df.isnull().all(axis=1)]

            file_cnt = 0
            # Loop through the indices and create new dataframes for each section
            for i in range(len(blank_row_indices)):
                if i == 0:
                    section_df = df.loc[:blank_row_indices[i]]
                else:
                    section_df = df.loc[blank_row_indices[i-1]+1:blank_row_indices[i]]
                
                # Create a new Excel file for the section
                section_df.to_excel(f'Section{i+1}.xlsx', index=False)
                file_cnt += 1
            #print(file_cnt)
            section_df = df.loc[blank_row_indices[-1]+1:]
            # Create a new Excel file for the section
            section_df.to_excel(f'Section{len(blank_row_indices)+1}.xlsx', index=False)


            for i in range(0,file_cnt+1):

                # Load the source workbook and sheet
                source_wb = openpyxl.load_workbook(f'Section{i+1}.xlsx')
                source_ws = source_wb['Sheet1']
                source_header = [cell.value for cell in source_ws[1]]

                # Load the destination workbook and sheet
                dest_wb = openpyxl.load_workbook('SHET_GC3-T1500_FM-XXX_Dive_XXX_EMPTY.xlsx')
                dest_ws = dest_wb['Data']
                dest_header = [cell.value for cell in dest_ws[2]]

                df = pd.DataFrame(list(source_ws.values))
                df.columns = df.iloc[0,:]
                df = df.iloc[1:,].reset_index(drop=True)

                df1 = pd.DataFrame(list(dest_ws.values))
                df1.columns = df1.iloc[1,:]
                df1 = df1.iloc[2:,].reset_index(drop=True)

                df['Port Fwd Load\n[bar]'] = df['Port Fwd Load\n[bar]'].astype(float)
                df['Stbd Fwd Load\n[bar]'] = df['Stbd Fwd Load\n[bar]'].astype(float)

                df['Average'] = (df['Port Fwd Load\n[bar]'] + df['Stbd Fwd Load\n[bar]'])/2

                df1['Tool Fwd/Aft Load \n[bar]'] = df['Average']

                df1[['Date\n[YYYY-MM-DD]', 'Time\n[HH:MM:SS]','KP\n[m]','DCC\n[m]','Depth\n[m]','Pitch\n[d.ddº]','Roll\n[d.ddº]','Port Sword\n[m]','Stbd Sword\n[m]','Jet Water Pressure\n[bar]','Comment']] = \
                df[["Date\n[YYYY-MM-DD]", "Time\n[HH:MM:SS]",'KP\n[km]','DCC\n[m]','Depth\n[m]','Pitch\n[d.ddº]','Roll\n[d.ddº]','Port Tip Depth\n[m]','Stbd Tip Depth\n[m]', 'Jet Water Pressure \n[bar]','Comment']]
                

                try:
                    df1[['Easting\n[m]','Northing\n[m]']]=df[['Easting\n[m]','Northing\n[m]','Easting\n[m]','Northing\n[m]']]
                    
                except:
                    df1[['Easting\n[m]', 'Northing\n[m]']] = df[['Easting\n[m]', 'Northing\n[m]']]
                
                
                 
                df1['Date\n[YYYY-MM-DD]'] = pd.to_datetime(df1['Date\n[YYYY-MM-DD]'])
                df1['Date\n[YYYY-MM-DD]'] = df1['Date\n[YYYY-MM-DD]'].dt.date
                
                # Delete blank rows in place
                df1 = df1.dropna(how='all')
                
                col_indices = [2,3,4]
                df1.iloc[:, col_indices] = None

                rows = dataframe_to_rows(df1, index=False)

                for r_idx, row in enumerate(rows, 2):
                    for c_idx, value in enumerate(row, 1):
                        dest_ws.cell(row=r_idx, column=c_idx, value=value)
                
                if (df_name['Comments'] == '1st Pass Trenching').any() and kp1 < kp2:
                    df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp1[i],3)}-KP{round(kp2[i],3)}_P1.xlsx', index=False)

                # Save the destination workbook
                    dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp1[i],3)}-KP{round(kp2[i],3)}_P1.xlsx')
                
                elif (df_name['Comments'] == '1st Pass Trenching').any() and kp1 > kp2:
                    df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp2[i],3)}-KP{round(kp1[i],3)}_P1.xlsx', index=False)

                # Save the destination workbook
                    dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp2[i],3)}-KP{round(kp1[i],3)}_P1.xlsx')
                
                elif (df_name['Comments'] == '2nd Pass Trenching').any() and kp1 < kp2: 
                
                    df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp1[i],3)}-KP{round(kp2[i],3)}_P2.xlsx', index=False)

                # Save the destination workbook
                    dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp1[i],3)}-KP{round(kp2[i],3)}_P2.xlsx')
                    
                elif (df_name['Comments'] == '2nd Pass Trenching').any() and kp1 > kp2: 
                
                    df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp2[i],3)}-KP{round(kp1[i],3)}_P2.xlsx', index=False)

                # Save the destination workbook
                    dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp2[i],3)}-KP{round(kp1[i],3)}_P2.xlsx')
                    
                elif (df_name['Comments'] == '3rd Pass Trenching').any() and kp1 < kp2:
                    df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp1[i],3)}-KP{round(kp2[i],3)}_P3.xlsx', index=False)

                # Save the destination workbook
                    dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp1[i],3)}-KP{round(kp2[i],3)}_P3.xlsx')
                    
                elif (df_name['Comments'] == '3rd Pass Trenching').any() and kp1 > kp2:
                    df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp2[i],3)}-KP{round(kp1[i],3)}_P3.xlsx', index=False)

                # Save the destination workbook
                    dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{round(kp2[i],3)}-KP{round(kp1[i],3)}_P3.xlsx')
                    
                
                
                
                os.remove(f'Section{i+1}.xlsx')
                #print("Script 1 executed")
        else:
            # Script 2: If the Excel file does not contain blank rows
            # Script 2 code here
            # ...
            # Load the source workbook and sheet
            
            rounded_kp1 = [round(num, 3) for num in kp1]
            rounded_kp2 = [round(num, 3) for num in kp2]
            
            source_wb = openpyxl.load_workbook(file)
            source_ws = source_wb['T1500']
            source_header = [cell.value for cell in source_ws[1]]

            # Load the destination workbook and sheet
            dest_wb = openpyxl.load_workbook('SHET_GC3-T1500_FM-XXX_Dive_XXX_EMPTY.xlsx')
            dest_ws = dest_wb['Data']
            dest_header = [cell.value for cell in dest_ws[2]]

            df = pd.DataFrame(list(source_ws.values))
            df.columns = df.iloc[0,:]
            df = df.iloc[1:,].reset_index(drop=True)

            df1 = pd.DataFrame(list(dest_ws.values))
            df1.columns = df1.iloc[1,:]
            df1 = df1.iloc[2:,].reset_index(drop=True)

            df['Port Fwd Load\n[bar]'] = df['Port Fwd Load\n[bar]'].astype(float)
            df['Stbd Fwd Load\n[bar]'] = df['Stbd Fwd Load\n[bar]'].astype(float)

            df['Average'] = (df['Port Fwd Load\n[bar]'] + df['Stbd Fwd Load\n[bar]'])/2

            df1['Tool Fwd/Aft Load \n[bar]'] = df['Average']

            df1[['Date\n[YYYY-MM-DD]', 'Time\n[HH:MM:SS]','KP\n[m]','DCC\n[m]','Depth\n[m]','Pitch\n[d.ddº]','Roll\n[d.ddº]','Port Sword\n[m]','Stbd Sword\n[m]','Jet Water Pressure\n[bar]','Comment']] = \
            df[["Date\n[YYYY-MM-DD]", "Time\n[HH:MM:SS]",'KP\n[km]','DCC\n[m]','Depth\n[m]','Pitch\n[d.ddº]','Roll\n[d.ddº]','Port Tip Depth\n[m]','Stbd Tip Depth\n[m]', 'Jet Water Pressure \n[bar]','Comment']]
                

            try:
                df1[['Easting\n[m]','Northing\n[m]']]=df[['Easting\n[m]','Northing\n[m]','Easting\n[m]','Northing\n[m]']]
                    
            except:
                df1[['Easting\n[m]', 'Northing\n[m]']] = df[['Easting\n[m]', 'Northing\n[m]']]
                
            
            df1['Date\n[YYYY-MM-DD]'] = pd.to_datetime(df1['Date\n[YYYY-MM-DD]'])
            df1['Date\n[YYYY-MM-DD]'] = df1['Date\n[YYYY-MM-DD]'].dt.date
            
            # Delete blank rows in place
            df1 = df1.dropna(how='all')    
            
            col_indices = [2,3,4]
            df1.iloc[:, col_indices] = None

            rows = dataframe_to_rows(df1, index=False)

            for r_idx, row in enumerate(rows, 2):
                for c_idx, value in enumerate(row, 1):
                    dest_ws.cell(row=r_idx, column=c_idx, value=value)
             
           
            if (df_name['Comments'] == '1st Pass Trenching').any() and kp1 < kp2:

                df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp1))}-KP{(", ".join(str(item) for item in rounded_kp2))}_P1.xlsx', index=False)

                # Save the destination workbook
                dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp1))}-KP{(", ".join(str(item) for item in rounded_kp2))}_P1.xlsx')
            
            elif (df_name['Comments'] == '1st Pass Trenching').any() and kp1 > kp2:

                df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp2))}-KP{(", ".join(str(item) for item in rounded_kp1))}_P1.xlsx', index=False)

                # Save the destination workbook
                dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp2))}-KP{(", ".join(str(item) for item in rounded_kp1))}_P1.xlsx')

            
                    
            elif (df_name['Comments'] == '2nd Pass Trenching').any() and kp1 < kp2:

                df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp1))}-KP{(", ".join(str(item) for item in rounded_kp2))}_P2.xlsx', index=False)

                # Save the destination workbook
                dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp1))}-KP{(", ".join(str(item) for item in rounded_kp2))}_P2.xlsx')
            
            elif (df_name['Comments'] == '2nd Pass Trenching').any() and kp1 > kp2:

                df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp2))}-KP{(", ".join(str(item) for item in rounded_kp1))}_P2.xlsx', index=False)

                # Save the destination workbook
                dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp2))}-KP{(", ".join(str(item) for item in rounded_kp1))}_P2.xlsx')
            
            
            elif (df_name['Comments'] == '3rd Pass Trenching').any() and kp1 < kp2:
                df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp1))}-KP{(", ".join(str(item) for item in rounded_kp2))}_P3.xlsx', index=False)

                # Save the destination workbook
                dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp1))}-KP{(", ".join(str(item) for item in rounded_kp2))}_P3.xlsx')
            
            elif (df_name['Comments'] == '3rd Pass Trenching').any() and kp1 > kp2:
                df1.to_excel(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp2))}-KP{(", ".join(str(item) for item in rounded_kp1))}_P3.xlsx', index=False)

                # Save the destination workbook
                dest_wb.save(f'SHET_GC3-T1500_FM-{FM_no}_Dive_{dive_no}_KP{(", ".join(str(item) for item in rounded_kp2))}-KP{(", ".join(str(item) for item in rounded_kp1))}_P3.xlsx')
            
            #print("Script 2 executed")

print("Script Completed !!!")